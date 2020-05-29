import openpyxl, os
os.chdir(r'C:\Users\zchen\OneDrive\Documents\Python stuff\Automate The Boring Stuff')
wb= openpyxl.Workbook() # write workbook
print(wb.get_sheet_names())
sheet= wb.get_sheet_by_name('Sheet')
sheet['A1']= 42 # writes cell
sheet2= wb.create_sheet()
sheet2.title= 'Randomsheetname'
sheet3= wb.create_sheet(index=0, title='1st sheet') # order the sheet and renaming title
print(wb.get_sheet_names())
wb.save('Myexcelfile.xlsx')