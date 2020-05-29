import openpyxl, os
os.chdir(r'C:\Users\zchen\OneDrive\Documents\Python stuff\Automate The Boring Stuff')
workbook= openpyxl.load_workbook('Calibration curve of standard.xlsx')
print(workbook.get_sheet_names())
sheet= workbook.get_sheet_by_name('Sheet1')
print(sheet['A4'].value) # Have to use .value to get the value in cell
# Another way to get value from cell
print(sheet.cell(row=5, column=2).value) # row and column start from 1

for i in range(1,10):
    print(i, sheet.cell(row=i, column=2).value) # all the values in the particular column
